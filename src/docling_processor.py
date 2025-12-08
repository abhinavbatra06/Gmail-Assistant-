"""
Docling Processor for Email Data

Process emails fetched by gmail_ingest.py using Docling
to convert EML files into structured JSON format ready for chunking.
"""

# CRITICAL: Set environment variables BEFORE any imports
import os
os.environ['HF_HUB_DISABLE_SYMLINKS_WARNING'] = '1'
os.environ['HF_HUB_DISABLE_SYMLINKS'] = '1'  # Fix Windows symlink permission error

# Fix SSL issues with HuggingFace using proper certificates
import ssl
import certifi
ssl._create_default_https_context = ssl.create_default_context(cafile=certifi.where())

import json
import yaml
from pathlib import Path
from typing import List, Dict, Optional
from bs4 import BeautifulSoup
from docling.document_converter import DocumentConverter
try:
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.datamodel.base_models import InputFormat
    from docling.document_converter import PdfFormatOption
except ImportError:
    # may not be available in all docling versions
    PdfPipelineOptions = None
    PdfFormatOption = None
from src.storage_manager import StorageManager
from src.db_helper import DBHelper
from src.ics_parser import ICSParser


class DoclingProcessor:

    def __init__(self, config_path="config.yaml"):
        with open(config_path, "r") as f:
            self.cfg = yaml.safe_load(f)

        self.paths = StorageManager(config_path)
        self.db = DBHelper(self.paths.db_path)
        self.docling_cfg = self.cfg.get("docling", {})
        
        # default initialization
        self.converter = DocumentConverter()

    def process_eml_file(self, eml_path: str, metadata: Optional[Dict] = None) -> Dict:
        """
        Process a single EML file using Docling.
        
        Args:
            eml_path: Path to the EML file
            metadata: Optional metadata dict from the metadata JSON file
            
        Returns:
            Dict containing structured document data ready for chunking
        """
        if not os.path.exists(eml_path):
            raise FileNotFoundError(f"EML file not found: {eml_path}")

        html_content = None
        text_content = None
        
        if metadata:
            html_content = metadata.get("body_html", "")
            text_content = metadata.get("body_text", "")
        
        # prefer using HTML content if available
        if html_content and html_content.strip():
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as tmp_file:
                tmp_file.write(html_content)
                tmp_path = tmp_file.name
            
            try:
                result = self.converter.convert(tmp_path)
                docling_doc = result.document
            finally:
                os.unlink(tmp_path)
        elif text_content and text_content.strip():
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as tmp_file:
                tmp_file.write(text_content)
                tmp_path = tmp_file.name
            
            try:
                result = self.converter.convert(tmp_path)
                docling_doc = result.document
            finally:
                os.unlink(tmp_path)
        else:
            result = self.converter.convert(eml_path)
            docling_doc = result.document

        extracted_text = docling_doc.text_content if hasattr(docling_doc, 'text_content') else ""
        
        if not extracted_text or not extracted_text.strip():
            if html_content and html_content.strip():
                soup = BeautifulSoup(html_content, 'html.parser')
                for script in soup(["script", "style"]):
                    script.decompose()
                extracted_text = soup.get_text(separator=' ', strip=True)
            elif text_content and text_content.strip():
                extracted_text = text_content
        
        structured_content = {
            "text": extracted_text,
            "tables": [],
            "metadata": {}
        }
        
        if hasattr(docling_doc, 'tables') and docling_doc.tables:
            for table in docling_doc.tables:
                table_data = {
                    "rows": len(table) if hasattr(table, '__len__') else 0,
                    "content": str(table) if hasattr(table, '__str__') else ""
                }
                structured_content["tables"].append(table_data)

        if metadata:
            structured_content["metadata"] = {
                "id": metadata.get("id"),
                "from": metadata.get("from"),
                "subject": metadata.get("subject"),
                "date": metadata.get("date"),
                "eml_path": eml_path,
                "attachments": metadata.get("attachments", [])
            }
        else:
            msg_id = Path(eml_path).stem
            structured_content["metadata"] = {
                "id": msg_id,
                "eml_path": eml_path
            }

        if hasattr(docling_doc, 'sections'):
            structured_content["sections"] = [
                {
                    "title": section.title if hasattr(section, 'title') else "",
                    "text": section.text if hasattr(section, 'text') else ""
                }
                for section in docling_doc.sections
            ]

        return structured_content

    def process_email(self, msg_id: str) -> Optional[Dict]:
        """
        Process a single email by its message ID.
        
        Args:
            msg_id: Gmail message ID
            
        Returns:
            Structured document dict or None if processing fails
        """
        try:
            eml_path = self.paths.path_for_email(msg_id)
            meta_path = self.paths.path_for_metadata(msg_id)
            docling_path = self.paths.path_for_docling(msg_id)

            if os.path.exists(docling_path):
                structured_doc = None
                try:
                    with open(docling_path, 'r', encoding='utf-8') as f:
                        structured_doc = json.load(f)
                except json.JSONDecodeError as json_err:
                    print(f"⚠️  Error processing JSON file for {msg_id}, will reprocess: {str(json_err)}")
                    # delete and reprocess
                    try:
                        os.remove(docling_path)
                    except:
                        pass
                    structured_doc = None
                except Exception as read_err:
                    print(f"⚠️  Error reading docling file for {msg_id}, will reprocess: {str(read_err)}")
                    # delete and reprocess
                    try:
                        if os.path.exists(docling_path):
                            os.remove(docling_path)
                    except:
                        pass
                    structured_doc = None
                
                if structured_doc is not None:
                    # check if attachments need to be processed
                    should_process_attachments = self.docling_cfg.get("process_attachments", True)
                    # check if attachments_processed key exists (indicates whether attachments were ever processed)
                    attachments_processed_key_exists = "attachments_processed" in structured_doc
                    existing_attachments = structured_doc.get("attachments_processed", [])
                    
                    # use database as source of truth for what attachments exist
                    # more reliable than metadata JSON which might be missing/out of sync
                    db_attachments = self.db.get_attachments_for_email(msg_id)
                    db_attachment_count = len(db_attachments) if db_attachments else 0
                    
                    # check processed attachment files on disk
                    processed_attachment_files = []
                    if db_attachment_count > 0:
                        for filename, _ in db_attachments:
                            processed_path = self.paths.path_for_processed_attachment(msg_id, filename)
                            if os.path.exists(processed_path):
                                processed_attachment_files.append(filename)

                    # determine if attachments need processing
                    has_unprocessed_attachments = False
                    if should_process_attachments and db_attachment_count > 0:
                        # if attachments_processed key doesn't exist, attachments were never processed
                        if not attachments_processed_key_exists:
                            has_unprocessed_attachments = True
                        # if database has more attachments than processed files on disk, need to process
                        elif db_attachment_count > len(processed_attachment_files):
                            has_unprocessed_attachments = True
                        # if attachments_processed exists but count doesn't match database, need to process
                        elif attachments_processed_key_exists and len(existing_attachments) < db_attachment_count:
                            has_unprocessed_attachments = True
                    
                    if has_unprocessed_attachments:
                        print(f"📎 Processing attachments for already-processed email: {msg_id} (DB has {db_attachment_count} attachments, {len(processed_attachment_files)} processed)")
                        attachments = self.process_attachments(msg_id, save=True)
                        structured_doc["attachments_processed"] = attachments
                        structured_doc["metadata"]["attachment_count"] = len(attachments)
                        
                        # ensure JSON-serializable before writing
                        structured_doc = self._ensure_json_serializable(structured_doc)
                        
                        # write to temp file first, then rename (atomic write)
                        temp_path = docling_path + ".tmp"
                        try:
                            with open(temp_path, 'w', encoding='utf-8') as f:
                                json.dump(structured_doc, f, ensure_ascii=False, indent=2)
                            # Atomic rename
                            os.replace(temp_path, docling_path)
                        except Exception as write_err:
                            # clean up temp file
                            if os.path.exists(temp_path):
                                try:
                                    os.remove(temp_path)
                                except:
                                    pass
                            raise write_err
                        
                        print(f"✅ Updated with attachments: {msg_id} | {structured_doc['metadata'].get('subject', '')[:60]}")
                    else:
                        print(f"⏭️  Already processed: {msg_id}")
                    
                    return structured_doc

            # email body hasn't been processed yet; process everything
            metadata = None
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                except (json.JSONDecodeError, Exception) as meta_err:
                    print(f"⚠️  Warning: Could not read metadata file for {msg_id}: {str(meta_err)}")
                    metadata = None  # continue without metadata

            structured_doc = self.process_eml_file(eml_path, metadata)

            if self.docling_cfg.get("process_attachments", True):
                attachments = self.process_attachments(msg_id, save=True)
                structured_doc["attachments_processed"] = attachments
                structured_doc["metadata"]["attachment_count"] = len(attachments)

            with open(docling_path, 'w', encoding='utf-8') as f:
                json.dump(structured_doc, f, ensure_ascii=False, indent=2)

            self.db.update_status(msg_id, "processed")

            print(f"Processed: {msg_id} | {structured_doc['metadata'].get('subject', '')[:60]}")
            return structured_doc

        except Exception as e:
            print(f" Error processing {msg_id}: {str(e)}")
            self.db.update_status(msg_id, f"error: {str(e)}")
            return None

    def process_all_emails(self, limit: Optional[int] = None, reprocess: bool = False) -> List[Dict]:
        """
        Process all emails that haven't been processed yet.
        
        Args:
            limit: Optional limit on number of emails to process
            reprocess: If True, also process emails that are already marked as 'processed'
                      (useful for processing attachments when config changes)
            
        Returns:
            List of processed document dicts
        """
        cur = self.db.conn.cursor()
        # always include processed emails if process_attachments is enabled
        should_process_attachments = self.docling_cfg.get("process_attachments", True)
        
        if reprocess or should_process_attachments:
            # include already processed emails to check for missing attachments
            query = "SELECT id FROM emails WHERE status = 'downloaded' OR status LIKE 'error:%' OR status = 'processed'"
        else:
            query = "SELECT id FROM emails WHERE status = 'downloaded' OR status LIKE 'error:%'"
        if limit:
            query += f" LIMIT {limit}"
        cur.execute(query)
        email_ids = [row[0] for row in cur.fetchall()]

        processed_docs = []
        for msg_id in email_ids:
            doc = self.process_email(msg_id)
            if doc:
                processed_docs.append(doc)

        return processed_docs

    def _ensure_json_serializable(self, obj):
        """
        Recursively ensure all objects in a dict/list are JSON serializable.
        Converts datetime objects, sets, and other non-serializable types.
        
        Args:
            obj: Object to make JSON-serializable
            
        Returns:
            JSON-serializable version of the object
        """
        from datetime import datetime, date
        
        if isinstance(obj, dict):
            return {k: self._ensure_json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._ensure_json_serializable(item) for item in obj]
        elif isinstance(obj, (datetime, date)):
            return obj.isoformat()
        elif isinstance(obj, set):
            return list(obj)
        elif hasattr(obj, '__dict__'):
            # convert custom objects to dict
            try:
                return self._ensure_json_serializable(obj.__dict__)
            except:
                return str(obj)
        else:
            # check if basic JSON-serializable type
            try:
                json.dumps(obj)
                return obj
            except (TypeError, ValueError):
                # fallback to string representation
                return str(obj)
    
    def _serialize_events(self, events: List[Dict]) -> List[Dict]:
        """
        Serialize event dictionaries for JSON storage.
        Converts datetime objects to ISO format strings.
        
        Args:
            events: List of event dictionaries with datetime objects
            
        Returns:
            List of event dictionaries with datetime objects converted to strings
        """
        from datetime import datetime
        
        serialized = []
        for event in events:
            serialized_event = event.copy()
            # convert datetime objects to ISO strings
            if isinstance(serialized_event.get("start_time"), datetime):
                serialized_event["start_time"] = serialized_event["start_time"].isoformat()
            if isinstance(serialized_event.get("end_time"), datetime):
                serialized_event["end_time"] = serialized_event["end_time"].isoformat()
            serialized.append(serialized_event)
        return serialized
    
    def _process_ics_file(self, att_path: str) -> str:
        """
        Process .ics (iCalendar) files by extracting text content.
        
        Args:
            att_path: Path to .ics file
            
        Returns:
            Extracted text content
        """
        try:
            with open(att_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            lines = content.split('\n')
            text_parts = []
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                if line.startswith(' '):
                    continue
                
                if line.startswith('BEGIN:'):
                    current_section = line.replace('BEGIN:', '').strip()
                    text_parts.append(f"\n[{current_section}]")
                elif line.startswith('END:'):
                    current_section = None
                elif ':' in line:
                    key, value = line.split(':', 1)
                    value = value.replace('\\n', '\n').replace('\\,', ',')
                    if key in ['SUMMARY', 'DESCRIPTION', 'LOCATION', 'DTSTART', 'DTEND', 'ORGANIZER']:
                        text_parts.append(f"{key}: {value}")
            
            return '\n'.join(text_parts)
        except Exception as e:
            return f"[ICS file - could not parse: {str(e)}]"

    def process_attachments(self, msg_id: str, save: bool = True) -> List[Dict]:
        """
        Process attachments for an email (PDFs, Word docs, ICS files, etc.).
        
        This method uses attachments that were already downloaded and saved by
        gmail_ingest.py. It reads attachment paths from the database (primary source)
        or metadata JSON file (fallback) and processes the files from disk.
        
        Args:
            msg_id: Gmail message ID
            save: Whether to save processed attachments to disk
            
        Returns:
            List of processed attachment documents
        """
        # get attachments from database (most reliable source)
        db_attachments = self.db.get_attachments_for_email(msg_id)
        
        # build attachments list from database
        attachments_meta = []
        if db_attachments:
            for filename, path in db_attachments:
                attachments_meta.append({"filename": filename, "path": path})
        
        # fallback to metadata JSON if database has no attachments
        if not attachments_meta:
            meta_path = self.paths.path_for_metadata(msg_id)
            if os.path.exists(meta_path):
                with open(meta_path, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                attachments_meta = metadata.get("attachments", [])
        if not attachments_meta:
            return []

        processed_attachments = []
        for att in attachments_meta:
            att_path = att.get("path")
            filename = att.get("filename", "")
            
            if not att_path or not os.path.exists(att_path):
                print(f"  Attachment file not found: {att_path} (may not have been downloaded)")
                continue

            ext = Path(att_path).suffix.lower()
            
            processed_path = self.paths.path_for_processed_attachment(msg_id, filename)
            if save and os.path.exists(processed_path):
                try:
                    with open(processed_path, 'r', encoding='utf-8') as f:
                        att_doc = json.load(f)
                    processed_attachments.append(att_doc)
                    continue
                except (json.JSONDecodeError, Exception) as read_err:
                    print(f"  ⚠️  Error decoding JSON for {filename}, will reprocess: {str(read_err)}")
                    # delete file and reprocess
                    try:
                        os.remove(processed_path)
                    except:
                        pass
                    # continue to process the attachment

            att_doc = {
                "filename": filename,
                "path": att_path,
                "text": "",
                "tables": [],
                "metadata": {
                    "email_id": msg_id,
                    "attachment_type": ext,
                    "original_filename": filename
                }
            }

            try:
                if ext == '.ics':
                    # use structured ICS parser
                    events = ICSParser.parse_ics_file(att_path)
                    if events:
                        # convert datetime objects to ISO strings for JSON serialization
                        serialized_events = self._serialize_events(events)
                        # store structured events
                        att_doc["events"] = serialized_events
                        # create text representation for backward compatibility
                        event_texts = [event["text"] for event in events]
                        att_doc["text"] = "\n\n---\n\n".join(event_texts)
                        att_doc["metadata"]["content_type"] = "calendar"
                        att_doc["metadata"]["event_count"] = len(events)
                        att_doc["metadata"]["is_calendar"] = True
                    else:
                        # fallback to old text extraction if parsing fails
                        att_doc["text"] = self._process_ics_file(att_path)
                        att_doc["metadata"]["content_type"] = "calendar"
                        att_doc["metadata"]["is_calendar"] = True
                    
                elif ext == '.xls':
                    # legacy Excel format; no docling support, use xlrd or openpyxl
                    try:
                        # try using xlrd for .xls files
                        try:
                            import xlrd
                            workbook = xlrd.open_workbook(att_path)
                            sheet_texts = []
                            for sheet in workbook.sheets():
                                sheet_texts.append(f"\n[Sheet: {sheet.name}]")
                                for row_idx in range(sheet.nrows):
                                    row_values = []
                                    for col_idx in range(sheet.ncols):
                                        cell_value = sheet.cell_value(row_idx, col_idx)
                                        if cell_value:
                                            row_values.append(str(cell_value))
                                    if row_values:
                                        sheet_texts.append(" | ".join(row_values))
                            extracted_text = "\n".join(sheet_texts)
                            att_doc["text"] = extracted_text
                            att_doc["metadata"]["content_type"] = "spreadsheet"
                            att_doc["metadata"]["file_format"] = "xls"
                        except ImportError:
                            try:
                                from openpyxl import load_workbook
                                workbook = load_workbook(att_path, read_only=True)
                                sheet_texts = []
                                for sheet_name in workbook.sheetnames:
                                    sheet = workbook[sheet_name]
                                    sheet_texts.append(f"\n[Sheet: {sheet_name}]")
                                    for row in sheet.iter_rows(values_only=True):
                                        row_values = [str(v) for v in row if v is not None]
                                        if row_values:
                                            sheet_texts.append(" | ".join(row_values))
                                extracted_text = "\n".join(sheet_texts)
                                att_doc["text"] = extracted_text
                                att_doc["metadata"]["content_type"] = "spreadsheet"
                                att_doc["metadata"]["file_format"] = "xls"
                            except Exception as e:
                                att_doc["text"] = f"[Excel file (.xls) - could not extract: {str(e)}. Install xlrd: pip install xlrd]"
                                att_doc["metadata"]["content_type"] = "spreadsheet"
                                att_doc["metadata"]["file_format"] = "xls"
                                att_doc["metadata"]["processing_error"] = str(e)
                    except Exception as e:
                        att_doc["text"] = f"[Excel file (.xls) - could not extract: {str(e)}]"
                        att_doc["metadata"]["content_type"] = "spreadsheet"
                        att_doc["metadata"]["file_format"] = "xls"
                        att_doc["metadata"]["processing_error"] = str(e)
                    
                elif ext in ['.pdf', '.docx', '.doc', '.pptx', '.xlsx']:
                    result = self.converter.convert(att_path)
                    docling_doc = result.document
                    
                    extracted_text = ""

                    if ext == '.xlsx':
                        # extract from tables first
                        if hasattr(docling_doc, 'tables') and docling_doc.tables:
                            table_texts = []
                            for table in docling_doc.tables:
                                if hasattr(table, 'data') and hasattr(table.data, 'table_cells'):
                                    # extract text from table cells
                                    cell_texts = []
                                    for cell in table.data.table_cells:
                                        if hasattr(cell, 'text') and cell.text:
                                            cell_texts.append(str(cell.text).strip())
                                    if cell_texts:
                                        table_texts.append(" | ".join(cell_texts))
                            if table_texts:
                                extracted_text = "\n".join(table_texts)
                    
                    if not extracted_text and hasattr(docling_doc, 'text_content'):
                        text_val = docling_doc.text_content
                        if text_val:
                            if callable(text_val):
                                try:
                                    extracted_text = text_val()
                                except:
                                    pass
                            else:
                                extracted_text = str(text_val) if text_val else ""
                    
                    if not extracted_text and hasattr(result, 'export_text'):
                        try:
                            extracted_text = result.export_text()
                        except Exception:
                            pass
                    
                    if not extracted_text and hasattr(docling_doc, 'texts'):
                        text_parts = []
                        try:
                            texts = docling_doc.texts
                            if texts:
                                for text_item in texts:
                                    if hasattr(text_item, 'text'):
                                        text_parts.append(str(text_item.text))
                                    elif isinstance(text_item, str):
                                        text_parts.append(text_item)
                                    else:
                                        text_parts.append(str(text_item))
                                extracted_text = "\n".join(filter(None, text_parts))
                        except Exception:
                            pass
                    
                    if not extracted_text:
                        try:
                            if hasattr(docling_doc, 'get_text'):
                                extracted_text = docling_doc.get_text()
                            elif hasattr(docling_doc, 'to_text'):
                                extracted_text = docling_doc.to_text()
                        except Exception:
                            pass
                    
                    att_doc["text"] = extracted_text
                    
                    if hasattr(docling_doc, 'tables') and docling_doc.tables:
                        for table in docling_doc.tables:
                            table_text = ""
                            if hasattr(table, 'data') and hasattr(table.data, 'table_cells'):
                                cell_texts = []
                                for cell in table.data.table_cells:
                                    if hasattr(cell, 'text') and cell.text:
                                        cell_texts.append(str(cell.text).strip())
                                if cell_texts:
                                    table_text = " | ".join(cell_texts)
                            else:
                                table_text = str(table)
                            
                            table_data = {
                                "rows": len(table) if hasattr(table, '__len__') else 0,
                                "content": table_text
                            }
                            att_doc["tables"].append(table_data)
                    
                    att_doc["metadata"]["content_type"] = "document" if ext != '.xlsx' else "spreadsheet"
                    if ext == '.xlsx':
                        att_doc["metadata"]["file_format"] = "xlsx"
                    
                    if not extracted_text:
                        file_type = "Excel spreadsheet" if ext == '.xlsx' else "document"
                        print(f"  Warning: No text extracted from {filename} (empty {file_type} or extraction failed)")
                    
                elif ext in ['.txt', '.csv', '.json', '.xml']:
                    with open(att_path, 'r', encoding='utf-8', errors='ignore') as f:
                        att_doc["text"] = f.read()
                    att_doc["metadata"]["content_type"] = "text"
                    
                elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg']:
                    try:
                        result = self.converter.convert(att_path)
                        docling_doc = result.document
                        
                        extracted_text = ""
                        
                        if hasattr(docling_doc, 'text_content'):
                            text_val = docling_doc.text_content
                            if text_val:
                                if callable(text_val):
                                    try:
                                        extracted_text = text_val()
                                    except:
                                        pass
                                else:
                                    extracted_text = str(text_val) if text_val else ""
                        
                        if not extracted_text and hasattr(result, 'export_text'):
                            try:
                                extracted_text = result.export_text()
                            except Exception:
                                pass
                        
                        if not extracted_text and hasattr(docling_doc, 'texts'):
                            text_parts = []
                            try:
                                texts = docling_doc.texts
                                if texts:
                                    for text_item in texts:
                                        if hasattr(text_item, 'text'):
                                            text_parts.append(str(text_item.text))
                                        elif isinstance(text_item, str):
                                            text_parts.append(text_item)
                                        else:
                                            text_parts.append(str(text_item))
                                    extracted_text = "\n".join(filter(None, text_parts))
                            except Exception:
                                pass
                        
                        if not extracted_text:
                            try:
                                if hasattr(docling_doc, 'get_text'):
                                    extracted_text = docling_doc.get_text()
                                elif hasattr(docling_doc, 'to_text'):
                                    extracted_text = docling_doc.to_text()
                            except Exception:
                                pass
                        
                        descriptions = []
                        if hasattr(docling_doc, 'pictures') and docling_doc.pictures:
                            for pic in docling_doc.pictures:
                                if hasattr(pic, 'description') and pic.description:
                                    descriptions.append(pic.description)
                        
                        if extracted_text:
                            att_doc["text"] = extracted_text
                        if descriptions:
                            if att_doc["text"]:
                                att_doc["text"] += "\n\n[Image Descriptions]\n"
                            else:
                                att_doc["text"] = "[Image Descriptions]\n"
                            att_doc["text"] += "\n".join(descriptions)
                        
                        if not att_doc["text"]:
                            att_doc["text"] = f"[Image file: {filename} - OCR did not extract text (may be image without text)]"
                            print(f"  Warning: No text extracted from image {filename}")
                        
                        att_doc["metadata"]["content_type"] = "image"
                        att_doc["metadata"]["image_format"] = ext
                        
                    except Exception as img_error:
                        att_doc["text"] = f"[Image file: {filename} - could not extract text: {str(img_error)}]"
                        att_doc["metadata"]["content_type"] = "image"
                        att_doc["metadata"]["image_format"] = ext
                        att_doc["metadata"]["processing_error"] = str(img_error)

                else:
                    print(f"  ⏭Skipping unsupported attachment type: {ext}")
                    continue

                if save:
                    # ensure JSON-serializable before writing
                    att_doc = self._ensure_json_serializable(att_doc)
                    
                    # write to temp file first, then rename (atomic write)
                    temp_path = processed_path + ".tmp"
                    try:
                        with open(temp_path, 'w', encoding='utf-8') as f:
                            json.dump(att_doc, f, ensure_ascii=False, indent=2)
                        # atomic rename
                        os.replace(temp_path, processed_path)
                    except Exception as write_err:
                        # clean up temp file 
                        if os.path.exists(temp_path):
                            try:
                                os.remove(temp_path)
                            except:
                                pass
                        raise write_err

                processed_attachments.append(att_doc)
                print(f"   Processed attachment: {filename} ({len(att_doc['text'])} chars)")

            except Exception as e:
                print(f"  Could not process attachment {filename}: {str(e)}")
                att_doc["metadata"]["error"] = str(e)
                if save:
                    # ensure JSON-serializable before writing
                    att_doc = self._ensure_json_serializable(att_doc)
                    
                    # write to temp file first, then rename (atomic write)
                    temp_path = processed_path + ".tmp"
                    try:
                        with open(temp_path, 'w', encoding='utf-8') as f:
                            json.dump(att_doc, f, ensure_ascii=False, indent=2)
                        # atomic rename
                        os.replace(temp_path, processed_path)
                    except Exception as write_err:
                        # clean up temp file if it exists
                        if os.path.exists(temp_path):
                            try:
                                os.remove(temp_path)
                            except:
                                pass
                        # don't raise; already logged the error

        return processed_attachments

    
    def get_processed_attachments(self, msg_id: str) -> List[Dict]:
        """
        Get all processed attachments for an email.
        
        Args:
            msg_id: Gmail message ID
            
        Returns:
            List of processed attachment documents
        """
        docling_path = self.paths.path_for_docling(msg_id)
        if os.path.exists(docling_path):
            with open(docling_path, 'r', encoding='utf-8') as f:
                doc = json.load(f)
            if doc.get("attachments_processed"):
                return doc["attachments_processed"]
        
        return self.process_attachments(msg_id, save=True)

    def close(self):
        """Close database connection."""
        self.db.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Process emails with Docling")
    parser.add_argument("--config", default="config.yaml", help="Path to config file")
    parser.add_argument("--limit", type=int, help="Limit number of emails to process")
    parser.add_argument("--msg-id", help="Process a specific message ID")
    parser.add_argument("--attachments", action="store_true", help="Also process attachments")

    args = parser.parse_args()

    print("Starting Docling Processor...")
    processor = DoclingProcessor(config_path=args.config)

    try:
        if args.msg_id:
            doc = processor.process_email(args.msg_id)
            if args.attachments:
                processor.process_attachments(args.msg_id)
        else:
            docs = processor.process_all_emails(limit=args.limit)
            print(f"\nDone. Processed {len(docs)} emails")
            if args.attachments:
                print("Processing attachments...")
                for doc in docs:
                    msg_id = doc["metadata"]["id"]
                    processor.process_attachments(msg_id)

    finally:
        processor.close()

